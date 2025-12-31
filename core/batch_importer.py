#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
批量导入模块
支持CSV和Excel文件批量导入目标数据
"""

import csv
import os
from pathlib import Path
from typing import List, Dict, Optional, Tuple, Any
from dataclasses import dataclass
from enum import Enum

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


class ImportFormat(Enum):
    """导入格式"""
    CSV = "csv"
    EXCEL = "excel"
    TSV = "tsv"
    TEXT = "text"


@dataclass
class TargetInfo:
    """目标信息"""
    url: str
    username: str
    password: str
    extra: Optional[Dict[str, str]] = None
    
    def is_valid(self) -> bool:
        """检查目标是否有效"""
        return bool(self.url and self.username and self.password)


@dataclass
class ImportResult:
    """导入结果"""
    success: bool
    targets: List[TargetInfo]
    total_rows: int
    valid_count: int
    invalid_count: int
    errors: List[str]
    warnings: List[str]


class BatchImporter:
    """批量导入器"""
    
    # 常见的URL字段名
    URL_FIELDS = ['url', 'URL', 'Url', '网址', '地址', 'address', 'target', '目标']
    
    # 常见的用户名字段名
    USERNAME_FIELDS = ['username', 'user', 'Username', 'User', '用户名', '用户', 'account', '账号', 'name']
    
    # 常见的密码字段名
    PASSWORD_FIELDS = ['password', 'pwd', 'Password', 'Pwd', '密码', 'pass', 'secret']
    
    def __init__(self):
        """初始化批量导入器"""
        self._check_dependencies()
    
    def _check_dependencies(self):
        """检查依赖"""
        self.openpyxl_available = OPENPYXL_AVAILABLE
        self.pandas_available = PANDAS_AVAILABLE
        
        if not self.openpyxl_available and not self.pandas_available:
            print("提示: 安装openpyxl或pandas可支持Excel导入: pip install openpyxl pandas")
    
    def import_file(self, filepath: str) -> ImportResult:
        """
        导入文件
        
        Args:
            filepath: 文件路径
            
        Returns:
            ImportResult
        """
        filepath = Path(filepath)
        
        if not filepath.exists():
            return ImportResult(
                success=False,
                targets=[],
                total_rows=0,
                valid_count=0,
                invalid_count=0,
                errors=[f"文件不存在: {filepath}"],
                warnings=[]
            )
        
        # 根据扩展名选择导入方法
        ext = filepath.suffix.lower()
        
        if ext == '.csv':
            return self.import_csv(str(filepath))
        elif ext in ['.xlsx', '.xls']:
            return self.import_excel(str(filepath))
        elif ext == '.tsv':
            return self.import_tsv(str(filepath))
        elif ext == '.txt':
            return self.import_text(str(filepath))
        else:
            return ImportResult(
                success=False,
                targets=[],
                total_rows=0,
                valid_count=0,
                invalid_count=0,
                errors=[f"不支持的文件格式: {ext}"],
                warnings=[]
            )
    
    def import_csv(self, filepath: str, encoding: str = 'utf-8-sig') -> ImportResult:
        """
        导入CSV文件
        
        Args:
            filepath: 文件路径
            encoding: 文件编码
            
        Returns:
            ImportResult
        """
        targets = []
        errors = []
        warnings = []
        total_rows = 0
        
        try:
            # 尝试多种编码
            encodings = [encoding, 'utf-8', 'gbk', 'gb2312', 'latin-1']
            content = None
            
            for enc in encodings:
                try:
                    with open(filepath, 'r', encoding=enc) as f:
                        content = f.read()
                        encoding = enc
                        break
                except UnicodeDecodeError:
                    continue
            
            if content is None:
                return ImportResult(
                    success=False,
                    targets=[],
                    total_rows=0,
                    valid_count=0,
                    invalid_count=0,
                    errors=["无法解码文件，请检查文件编码"],
                    warnings=[]
                )
            
            # 解析CSV
            reader = csv.DictReader(content.splitlines())
            
            # 识别字段映射
            fieldnames = reader.fieldnames or []
            url_field = self._find_field(fieldnames, self.URL_FIELDS)
            username_field = self._find_field(fieldnames, self.USERNAME_FIELDS)
            password_field = self._find_field(fieldnames, self.PASSWORD_FIELDS)
            
            if not all([url_field, username_field, password_field]):
                missing = []
                if not url_field:
                    missing.append("URL")
                if not username_field:
                    missing.append("用户名")
                if not password_field:
                    missing.append("密码")
                
                return ImportResult(
                    success=False,
                    targets=[],
                    total_rows=0,
                    valid_count=0,
                    invalid_count=0,
                    errors=[f"缺少必需字段: {', '.join(missing)}。可用字段: {', '.join(fieldnames)}"],
                    warnings=[]
                )
            
            # 读取数据
            for row_num, row in enumerate(reader, start=2):
                total_rows += 1
                
                try:
                    url = row.get(url_field, '').strip()
                    username = row.get(username_field, '').strip()
                    password = row.get(password_field, '').strip()
                    
                    # 收集额外字段
                    extra = {}
                    for key, value in row.items():
                        if key not in [url_field, username_field, password_field]:
                            extra[key] = value.strip() if value else ''
                    
                    target = TargetInfo(
                        url=url,
                        username=username,
                        password=password,
                        extra=extra if extra else None
                    )
                    
                    if target.is_valid():
                        targets.append(target)
                    else:
                        warnings.append(f"第{row_num}行数据不完整，已跳过")
                        
                except Exception as e:
                    errors.append(f"第{row_num}行解析失败: {str(e)}")
            
            return ImportResult(
                success=len(targets) > 0,
                targets=targets,
                total_rows=total_rows,
                valid_count=len(targets),
                invalid_count=total_rows - len(targets),
                errors=errors,
                warnings=warnings
            )
            
        except Exception as e:
            return ImportResult(
                success=False,
                targets=[],
                total_rows=0,
                valid_count=0,
                invalid_count=0,
                errors=[f"CSV导入失败: {str(e)}"],
                warnings=[]
            )
    
    def import_excel(self, filepath: str) -> ImportResult:
        """
        导入Excel文件
        
        Args:
            filepath: 文件路径
            
        Returns:
            ImportResult
        """
        # 优先使用pandas（更强大）
        if self.pandas_available:
            return self._import_excel_pandas(filepath)
        elif self.openpyxl_available:
            return self._import_excel_openpyxl(filepath)
        else:
            return ImportResult(
                success=False,
                targets=[],
                total_rows=0,
                valid_count=0,
                invalid_count=0,
                errors=["Excel导入需要安装openpyxl或pandas: pip install openpyxl pandas"],
                warnings=[]
            )
    
    def _import_excel_pandas(self, filepath: str) -> ImportResult:
        """使用pandas导入Excel"""
        targets = []
        errors = []
        warnings = []
        total_rows = 0
        
        try:
            df = pd.read_excel(filepath)
            total_rows = len(df)
            
            # 识别字段映射
            columns = df.columns.tolist()
            url_field = self._find_field(columns, self.URL_FIELDS)
            username_field = self._find_field(columns, self.USERNAME_FIELDS)
            password_field = self._find_field(columns, self.PASSWORD_FIELDS)
            
            if not all([url_field, username_field, password_field]):
                missing = []
                if not url_field:
                    missing.append("URL")
                if not username_field:
                    missing.append("用户名")
                if not password_field:
                    missing.append("密码")
                
                return ImportResult(
                    success=False,
                    targets=[],
                    total_rows=total_rows,
                    valid_count=0,
                    invalid_count=total_rows,
                    errors=[f"缺少必需字段: {', '.join(missing)}。可用字段: {', '.join(columns)}"],
                    warnings=[]
                )
            
            # 读取数据
            for idx, row in df.iterrows():
                row_num = idx + 2  # Excel行号从2开始（1是标题）
                
                try:
                    url = str(row.get(url_field, '')).strip()
                    username = str(row.get(username_field, '')).strip()
                    password = str(row.get(password_field, '')).strip()
                    
                    # 处理NaN值
                    if url == 'nan':
                        url = ''
                    if username == 'nan':
                        username = ''
                    if password == 'nan':
                        password = ''
                    
                    # 收集额外字段
                    extra = {}
                    for col in columns:
                        if col not in [url_field, username_field, password_field]:
                            val = str(row.get(col, '')).strip()
                            if val and val != 'nan':
                                extra[col] = val
                    
                    target = TargetInfo(
                        url=url,
                        username=username,
                        password=password,
                        extra=extra if extra else None
                    )
                    
                    if target.is_valid():
                        targets.append(target)
                    else:
                        warnings.append(f"第{row_num}行数据不完整，已跳过")
                        
                except Exception as e:
                    errors.append(f"第{row_num}行解析失败: {str(e)}")
            
            return ImportResult(
                success=len(targets) > 0,
                targets=targets,
                total_rows=total_rows,
                valid_count=len(targets),
                invalid_count=total_rows - len(targets),
                errors=errors,
                warnings=warnings
            )
            
        except Exception as e:
            return ImportResult(
                success=False,
                targets=[],
                total_rows=0,
                valid_count=0,
                invalid_count=0,
                errors=[f"Excel导入失败: {str(e)}"],
                warnings=[]
            )
    
    def _import_excel_openpyxl(self, filepath: str) -> ImportResult:
        """使用openpyxl导入Excel"""
        targets = []
        errors = []
        warnings = []
        total_rows = 0
        
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            
            # 获取标题行
            headers = [cell.value for cell in ws[1] if cell.value]
            
            # 识别字段映射
            url_idx = self._find_field_index(headers, self.URL_FIELDS)
            username_idx = self._find_field_index(headers, self.USERNAME_FIELDS)
            password_idx = self._find_field_index(headers, self.PASSWORD_FIELDS)
            
            if url_idx is None or username_idx is None or password_idx is None:
                missing = []
                if url_idx is None:
                    missing.append("URL")
                if username_idx is None:
                    missing.append("用户名")
                if password_idx is None:
                    missing.append("密码")
                
                return ImportResult(
                    success=False,
                    targets=[],
                    total_rows=0,
                    valid_count=0,
                    invalid_count=0,
                    errors=[f"缺少必需字段: {', '.join(missing)}。可用字段: {', '.join(headers)}"],
                    warnings=[]
                )
            
            # 读取数据行
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # 跳过空行
                    continue
                    
                total_rows += 1
                
                try:
                    url = str(row[url_idx] or '').strip()
                    username = str(row[username_idx] or '').strip()
                    password = str(row[password_idx] or '').strip()
                    
                    # 收集额外字段
                    extra = {}
                    for i, header in enumerate(headers):
                        if i not in [url_idx, username_idx, password_idx] and i < len(row):
                            val = str(row[i] or '').strip()
                            if val:
                                extra[header] = val
                    
                    target = TargetInfo(
                        url=url,
                        username=username,
                        password=password,
                        extra=extra if extra else None
                    )
                    
                    if target.is_valid():
                        targets.append(target)
                    else:
                        warnings.append(f"第{row_num}行数据不完整，已跳过")
                        
                except Exception as e:
                    errors.append(f"第{row_num}行解析失败: {str(e)}")
            
            wb.close()
            
            return ImportResult(
                success=len(targets) > 0,
                targets=targets,
                total_rows=total_rows,
                valid_count=len(targets),
                invalid_count=total_rows - len(targets),
                errors=errors,
                warnings=warnings
            )
            
        except Exception as e:
            return ImportResult(
                success=False,
                targets=[],
                total_rows=0,
                valid_count=0,
                invalid_count=0,
                errors=[f"Excel导入失败: {str(e)}"],
                warnings=[]
            )
    
    def import_tsv(self, filepath: str) -> ImportResult:
        """
        导入TSV文件
        
        Args:
            filepath: 文件路径
            
        Returns:
            ImportResult
        """
        targets = []
        errors = []
        warnings = []
        total_rows = 0
        
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f, delimiter='\t')
                
                fieldnames = reader.fieldnames or []
                url_field = self._find_field(fieldnames, self.URL_FIELDS)
                username_field = self._find_field(fieldnames, self.USERNAME_FIELDS)
                password_field = self._find_field(fieldnames, self.PASSWORD_FIELDS)
                
                if not all([url_field, username_field, password_field]):
                    return ImportResult(
                        success=False,
                        targets=[],
                        total_rows=0,
                        valid_count=0,
                        invalid_count=0,
                        errors=["缺少必需字段（URL、用户名、密码）"],
                        warnings=[]
                    )
                
                for row_num, row in enumerate(reader, start=2):
                    total_rows += 1
                    
                    try:
                        target = TargetInfo(
                            url=row.get(url_field, '').strip(),
                            username=row.get(username_field, '').strip(),
                            password=row.get(password_field, '').strip()
                        )
                        
                        if target.is_valid():
                            targets.append(target)
                        else:
                            warnings.append(f"第{row_num}行数据不完整")
                            
                    except Exception as e:
                        errors.append(f"第{row_num}行解析失败: {str(e)}")
                
            return ImportResult(
                success=len(targets) > 0,
                targets=targets,
                total_rows=total_rows,
                valid_count=len(targets),
                invalid_count=total_rows - len(targets),
                errors=errors,
                warnings=warnings
            )
            
        except Exception as e:
            return ImportResult(
                success=False,
                targets=[],
                total_rows=0,
                valid_count=0,
                invalid_count=0,
                errors=[f"TSV导入失败: {str(e)}"],
                warnings=[]
            )
    
    def import_text(self, filepath: str) -> ImportResult:
        """
        导入文本文件（每行格式: URL,用户名,密码）
        
        Args:
            filepath: 文件路径
            
        Returns:
            ImportResult
        """
        targets = []
        errors = []
        warnings = []
        total_rows = 0
        
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                lines = f.readlines()
            
            for row_num, line in enumerate(lines, start=1):
                line = line.strip()
                if not line or line.startswith('#'):  # 跳过空行和注释
                    continue
                
                total_rows += 1
                
                try:
                    # 尝试不同的分隔符
                    parts = None
                    for sep in [',', '\t', '|', ' ']:
                        if sep in line:
                            parts = line.split(sep)
                            break
                    
                    if not parts or len(parts) < 3:
                        warnings.append(f"第{row_num}行格式错误，需要: URL,用户名,密码")
                        continue
                    
                    target = TargetInfo(
                        url=parts[0].strip(),
                        username=parts[1].strip(),
                        password=parts[2].strip()
                    )
                    
                    if target.is_valid():
                        targets.append(target)
                    else:
                        warnings.append(f"第{row_num}行数据不完整")
                        
                except Exception as e:
                    errors.append(f"第{row_num}行解析失败: {str(e)}")
            
            return ImportResult(
                success=len(targets) > 0,
                targets=targets,
                total_rows=total_rows,
                valid_count=len(targets),
                invalid_count=total_rows - len(targets),
                errors=errors,
                warnings=warnings
            )
            
        except Exception as e:
            return ImportResult(
                success=False,
                targets=[],
                total_rows=0,
                valid_count=0,
                invalid_count=0,
                errors=[f"文本文件导入失败: {str(e)}"],
                warnings=[]
            )
    
    def _find_field(self, fields: List[str], candidates: List[str]) -> Optional[str]:
        """查找匹配的字段名"""
        for candidate in candidates:
            for field in fields:
                if field.lower() == candidate.lower():
                    return field
        return None
    
    def _find_field_index(self, fields: List[str], candidates: List[str]) -> Optional[int]:
        """查找匹配的字段索引"""
        for candidate in candidates:
            for i, field in enumerate(fields):
                if field and field.lower() == candidate.lower():
                    return i
        return None
    
    def export_template(self, filepath: str, format: ImportFormat = ImportFormat.CSV) -> bool:
        """
        导出模板文件
        
        Args:
            filepath: 文件路径
            format: 导出格式
            
        Returns:
            是否成功
        """
        try:
            if format == ImportFormat.CSV:
                with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f)
                    writer.writerow(['URL', '用户名', '密码'])
                    writer.writerow(['https://example.com/login', 'admin', 'password123'])
                    writer.writerow(['https://example.org/auth', 'user1', 'pass456'])
                return True
                
            elif format == ImportFormat.EXCEL and self.openpyxl_available:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "目标列表"
                ws.append(['URL', '用户名', '密码'])
                ws.append(['https://example.com/login', 'admin', 'password123'])
                ws.append(['https://example.org/auth', 'user1', 'pass456'])
                wb.save(filepath)
                wb.close()
                return True
                
            else:
                return False
                
        except Exception as e:
            print(f"导出模板失败: {e}")
            return False


def get_supported_formats() -> List[str]:
    """
    获取支持的导入格式
    
    Returns:
        格式列表
    """
    formats = ['.csv', '.tsv', '.txt']
    
    if OPENPYXL_AVAILABLE or PANDAS_AVAILABLE:
        formats.extend(['.xlsx', '.xls'])
    
    return formats


if __name__ == "__main__":
    # 测试代码
    print("支持的导入格式:", get_supported_formats())
    
    importer = BatchImporter()
    
    # 创建测试CSV文件
    test_csv = "test_import.csv"
    with open(test_csv, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(['URL', '用户名', '密码'])
        writer.writerow(['https://httpbin.org/post', 'testuser1', 'testpass1'])
        writer.writerow(['https://httpbin.org/post', 'testuser2', 'testpass2'])
    
    # 测试导入
    result = importer.import_csv(test_csv)
    print(f"\n导入结果:")
    print(f"  成功: {result.success}")
    print(f"  总行数: {result.total_rows}")
    print(f"  有效数: {result.valid_count}")
    print(f"  无效数: {result.invalid_count}")
    print(f"  目标列表:")
    for target in result.targets:
        print(f"    - {target.url} | {target.username}")
    
    # 清理测试文件
    os.remove(test_csv)