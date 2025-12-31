#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
增强版批量导入模块
支持CSV、Excel、TSV、TXT文件批量导入目标数据
添加类型注解和错误处理优化
"""

import csv
import os
import json
from pathlib import Path
from typing import List, Dict, Optional, Tuple, Any, Union, Set
from dataclasses import dataclass, field
from enum import Enum
import io

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


class ImportFormat(Enum):
    """导入文件格式"""
    CSV = "csv"
    EXCEL = "excel"
    TSV = "tsv"
    TEXT = "text"
    JSON = "json"


@dataclass
class TargetInfo:
    """目标信息"""
    url: str
    username: str
    password: str
    extra: Optional[Dict[str, str]] = field(default=None)
    index: int = 0
    source_file: Optional[str] = field(default=None)
    
    def is_valid(self) -> bool:
        """检查目标是否有效"""
        return bool(self.url and self.username and self.password)
    
    def to_dict(self) -> Dict[str, Any]:
        """转换为字典"""
        data = {
            'url': self.url,
            'username': self.username,
            'password': self.password,
            'index': self.index
        }
        if self.extra:
            data['extra'] = self.extra
        if self.source_file:
            data['source_file'] = self.source_file
        return data


@dataclass
class ImportResult:
    """导入结果"""
    success: bool
    targets: List[TargetInfo]
    total_rows: int
    valid_count: int
    invalid_count: int
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    format: Optional[ImportFormat] = field(default=None)
    encoding: str = "utf-8"


class EnhancedBatchImporter:
    """增强版批量导入器"""
    
    def __init__(self):
        """初始化批量导入器"""
        self._check_dependencies()
        
        # URL字段名（支持中英文）
        self.URL_FIELDS: Set[str] = {
            'url', 'URL', 'Url', '网址', '地址', 'address', 'target', '目标',
            'host', '主机', '链接', 'link', 'site', '站点'
        }
        
        # 用户名字段名
        self.USERNAME_FIELDS: Set[str] = {
            'username', 'user', 'Username', 'User', '用户名', '用户',
            'account', '账号', 'name', 'login', '登录名', 'uid', 'userid',
            'mobile', 'phone', 'tel', 'email', 'mail'
        }
        
        # 密码字段名
        self.PASSWORD_FIELDS: Set[str] = {
            'password', 'pwd', 'pass', 'Password', 'Pwd', 'Pass', '密码',
            'passwd', 'secret', 'key', 'passcode', '口令'
        }
    
    def _check_dependencies(self):
        """检查依赖"""
        self.openpyxl_available = OPENPYXL_AVAILABLE
        self.pandas_available = PANDAS_AVAILABLE
        
        if not self.openpyxl_available and not self.pandas_available:
            print("提示: 安装openpyxl或pandas可支持Excel导入")
    
    def detect_encoding(self, filepath: str) -> str:
        """
        检测文件编码
        
        Args:
            filepath: 文件路径
            
        Returns:
            检测到的编码
        """
        encodings = ['utf-8', 'utf-8-sig', 'gbk', 'gb2312', 'gb18030', 'latin-1']
        
        for encoding in encodings:
            try:
                with open(filepath, 'r', encoding=encoding) as f:
                    f.read()
                return encoding
            except (UnicodeDecodeError, UnicodeError):
                continue
        
        return 'utf-8'
    
    def detect_format(self, filepath: str) -> Optional[ImportFormat]:
        """
        检测文件格式
        
        Args:
            filepath: 文件路径
            
        Returns:
            文件格式
        """
        ext = Path(filepath).suffix.lower()
        
        format_map = {
            '.csv': ImportFormat.CSV,
            '.xlsx': ImportFormat.EXCEL,
            '.xls': ImportFormat.EXCEL,
            '.tsv': ImportFormat.TSV,
            '.txt': ImportFormat.TEXT,
            '.json': ImportFormat.JSON
        }
        
        return format_map.get(ext)
    
    def import_file(self, filepath: str) -> ImportResult:
        """
        导入文件（自动检测格式）
        
        Args:
            filepath: 文件路径
            
        Returns:
            ImportResult
        """
        filepath = Path(filepath)
        
        if not filepath.exists():
            return ImportResult(
                success=False,
                targets=[],
                total_rows=0,
                valid_count=0,
                invalid_count=0,
                errors=[f"文件不存在: {filepath}"]
            )
        
        file_format = self.detect_format(str(filepath))
        
        if not file_format:
            return ImportResult(
                success=False,
                targets=[],
                total_rows=0,
                valid_count=0,
                invalid_count=0,
                errors=[f"不支持的文件格式: {filepath.suffix}"]
            )
        
        if file_format == ImportFormat.CSV:
            return self.import_csv(str(filepath))
        elif file_format == ImportFormat.EXCEL:
            return self.import_excel(str(filepath))
        elif file_format == ImportFormat.TSV:
            return self.import_tsv(str(filepath))
        elif file_format == ImportFormat.TEXT:
            return self.import_text(str(filepath))
        elif file_format == ImportFormat.JSON:
            return self.import_json(str(filepath))
        else:
            return ImportResult(
                success=False,
                targets=[],
                total_rows=0,
                valid_count=0,
                invalid_count=0,
                errors=[f"未实现的格式: {file_format.value}"]
            )
    
    def import_csv(self, filepath: str, encoding: Optional[str] = None) -> ImportResult:
        """
        导入CSV文件
        
        Args:
            filepath: 文件路径
            encoding: 文件编码（自动检测）
            
        Returns:
            ImportResult
        """
        targets: List[TargetInfo] = []
        errors: List[str] = []
        warnings: List[str] = []
        total_rows = 0
        
        try:
            if encoding is None:
                encoding = self.detect_encoding(filepath)
            
            with open(filepath, 'r', encoding=encoding) as f:
                content = f.read()
            
            reader = csv.DictReader(content.splitlines())
            
            if not reader.fieldnames:
                return ImportResult(
                    success=False,
                    targets=[],
                    total_rows=0,
                    valid_count=0,
                    invalid_count=0,
                    errors=["CSV文件为空或格式错误"]
                )
            
            fieldnames = reader.fieldnames or []
            url_field = self._find_field(fieldnames, self.URL_FIELDS)
            username_field = self._find_field(fieldnames, self.USERNAME_FIELDS)
            password_field = self._find_field(fieldnames, self.PASSWORD_FIELDS)
            
            if not all([url_field, username_field, password_field]):
                missing = []
                if not url_field:
                    missing.append("URL")
                if not username_field:
                    missing.append("用户名")
                if not password_field:
                    missing.append("密码")
                
                return ImportResult(
                    success=False,
                    targets=[],
                    total_rows=0,
                    valid_count=0,
                    invalid_count=0,
                    errors=[f"缺少必需字段: {', '.join(missing)}。可用字段: {', '.join(fieldnames)}"]
                )
            
            for row_num, row in enumerate(reader, start=2):
                total_rows += 1
                
                try:
                    # 调试输出
                    # print(f"DEBUG: 处理第{row_num}行, row类型: {type(row)}, row内容: {row}")
                    
                    url = row.get(url_field, '').strip()
                    username = row.get(username_field, '').strip()
                    password = row.get(password_field, '').strip()
                    
                    extra: Dict[str, str] = {}
                    for key, value in row.items():
                        if key not in [url_field, username_field, password_field] and value:
                            # 确保value是字符串
                            value_str = str(value) if not isinstance(value, str) else value
                            extra[key] = value_str.strip()
                    
                    target = TargetInfo(
                        url=url,
                        username=username,
                        password=password,
                        extra=extra if extra else None,
                        index=len(targets),
                        source_file=Path(filepath).name
                    )
                    
                    if target.is_valid():
                        targets.append(target)
                    else:
                        warnings.append(f"第{row_num}行数据不完整，已跳过")
                        
                except Exception as e:
                    errors.append(f"第{row_num}行解析失败: {str(e)}")
                    import traceback
                    # 打印完整的traceback用于调试
                    # print(f"DEBUG: 第{row_num}行异常详情: {traceback.format_exc()}")
            
            return ImportResult(
                success=len(targets) > 0,
                targets=targets,
                total_rows=total_rows,
                valid_count=len(targets),
                invalid_count=total_rows - len(targets),
                errors=errors,
                warnings=warnings,
                format=ImportFormat.CSV,
                encoding=encoding
            )
            
        except Exception as e:
            return ImportResult(
                success=False,
                targets=[],
                total_rows=0,
                valid_count=0,
                invalid_count=0,
                errors=[f"CSV导入失败: {str(e)}"],
                format=ImportFormat.CSV
            )
    
    def import_excel(self, filepath: str) -> ImportResult:
        """
        导入Excel文件
        
        Args:
            filepath: 文件路径
            
        Returns:
            ImportResult
        """
        if self.pandas_available:
            return self._import_excel_pandas(filepath)
        elif self.openpyxl_available:
            return self._import_excel_openpyxl(filepath)
        else:
            return ImportResult(
                success=False,
                targets=[],
                total_rows=0,
                valid_count=0,
                invalid_count=0,
                errors=["Excel导入需要安装openpyxl或pandas"],
                format=ImportFormat.EXCEL
            )
    
    def _import_excel_pandas(self, filepath: str) -> ImportResult:
        """使用pandas导入Excel"""
        targets: List[TargetInfo] = []
        errors: List[str] = []
        warnings: List[str] = []
        total_rows = 0
        
        try:
            df = pd.read_excel(filepath)
            total_rows = len(df)
            
            columns = [str(col).strip() for col in df.columns]
            url_field = self._find_field(columns, self.URL_FIELDS)
            username_field = self._find_field(columns, self.USERNAME_FIELDS)
            password_field = self._find_field(columns, self.PASSWORD_FIELDS)
            
            if not all([url_field, username_field, password_field]):
                missing = []
                if not url_field:
                    missing.append("URL")
                if not username_field:
                    missing.append("用户名")
                if not password_field:
                    missing.append("密码")
                
                return ImportResult(
                    success=False,
                    targets=[],
                    total_rows=total_rows,
                    valid_count=0,
                    invalid_count=total_rows,
                    errors=[f"缺少必需字段: {', '.join(missing)}"],
                    format=ImportFormat.EXCEL
                )
            
            for idx, row in df.iterrows():
                row_num = idx + 2
                
                try:
                    url = str(row.get(url_field, '')).strip()
                    username = str(row.get(username_field, '')).strip()
                    password = str(row.get(password_field, '')).strip()
                    
                    if url in ['nan', 'NaN', '']:
                        url = ''
                    if username in ['nan', 'NaN', '']:
                        username = ''
                    if password in ['nan', 'NaN', '']:
                        password = ''
                    
                    extra: Dict[str, str] = {}
                    for col in columns:
                        if col not in [url_field, username_field, password_field]:
                            val = str(row.get(col, '')).strip()
                            if val and val not in ['nan', 'NaN', '']:
                                extra[col] = val
                    
                    target = TargetInfo(
                        url=url,
                        username=username,
                        password=password,
                        extra=extra if extra else None,
                        index=len(targets),
                        source_file=Path(filepath).name
                    )
                    
                    if target.is_valid():
                        targets.append(target)
                    else:
                        warnings.append(f"第{row_num}行数据不完整，已跳过")
                        
                except Exception as e:
                    errors.append(f"第{row_num}行解析失败: {str(e)}")
            
            return ImportResult(
                success=len(targets) > 0,
                targets=targets,
                total_rows=total_rows,
                valid_count=len(targets),
                invalid_count=total_rows - len(targets),
                errors=errors,
                warnings=warnings,
                format=ImportFormat.EXCEL
            )
            
        except Exception as e:
            return ImportResult(
                success=False,
                targets=[],
                total_rows=0,
                valid_count=0,
                invalid_count=0,
                errors=[f"Excel导入失败: {str(e)}"],
                format=ImportFormat.EXCEL
            )
    
    def _import_excel_openpyxl(self, filepath: str) -> ImportResult:
        """使用openpyxl导入Excel"""
        targets: List[TargetInfo] = []
        errors: List[str] = []
        warnings: List[str] = []
        total_rows = 0
        
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            
            headers = [str(cell.value or '') for cell in ws[1] if cell.value]
            
            url_idx = self._find_field_index(headers, self.URL_FIELDS)
            username_idx = self._find_field_index(headers, self.USERNAME_FIELDS)
            password_idx = self._find_field_index(headers, self.PASSWORD_FIELDS)
            
            if url_idx is None or username_idx is None or password_idx is None:
                missing = []
                if url_idx is None:
                    missing.append("URL")
                if username_idx is None:
                    missing.append("用户名")
                if password_idx is None:
                    missing.append("密码")
                
                return ImportResult(
                    success=False,
                    targets=[],
                    total_rows=0,
                    valid_count=0,
                    invalid_count=0,
                    errors=[f"缺少必需字段: {', '.join(missing)}"]
                )
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue
                    
                total_rows += 1
                
                try:
                    url = str(row[url_idx] or '').strip()
                    username = str(row[username_idx] or '').strip()
                    password = str(row[password_idx] or '').strip()
                    
                    extra: Dict[str, str] = {}
                    for i, header in enumerate(headers):
                        if i not in [url_idx, username_idx, password_idx] and i < len(row):
                            val = str(row[i] or '').strip()
                            if val:
                                extra[header] = val
                    
                    target = TargetInfo(
                        url=url,
                        username=username,
                        password=password,
                        extra=extra if extra else None,
                        index=len(targets),
                        source_file=Path(filepath).name
                    )
                    
                    if target.is_valid():
                        targets.append(target)
                    else:
                        warnings.append(f"第{row_num}行数据不完整，已跳过")
                        
                except Exception as e:
                    errors.append(f"第{row_num}行解析失败: {str(e)}")
            
            wb.close()
            
            return ImportResult(
                success=len(targets) > 0,
                targets=targets,
                total_rows=total_rows,
                valid_count=len(targets),
                invalid_count=total_rows - len(targets),
                errors=errors,
                warnings=warnings,
                format=ImportFormat.EXCEL
            )
            
        except Exception as e:
            return ImportResult(
                success=False,
                targets=[],
                total_rows=0,
                valid_count=0,
                invalid_count=0,
                errors=[f"Excel导入失败: {str(e)}"],
                format=ImportFormat.EXCEL
            )
    
    def import_tsv(self, filepath: str) -> ImportResult:
        """导入TSV文件"""
        targets: List[TargetInfo] = []
        errors: List[str] = []
        warnings: List[str] = []
        total_rows = 0
        
        try:
            encoding = self.detect_encoding(filepath)
            
            with open(filepath, 'r', encoding=encoding) as f:
                reader = csv.DictReader(f, delimiter='\t')
                
                fieldnames = reader.fieldnames or []
                url_field = self._find_field(fieldnames, self.URL_FIELDS)
                username_field = self._find_field(fieldnames, self.USERNAME_FIELDS)
                password_field = self._find_field(fieldnames, self.PASSWORD_FIELDS)
                
                if not all([url_field, username_field, password_field]):
                    return ImportResult(
                        success=False,
                        targets=[],
                        total_rows=0,
                        valid_count=0,
                        invalid_count=0,
                        errors=["缺少必需字段"]
                    )
                
                for row_num, row in enumerate(reader, start=2):
                    total_rows += 1
                    
                    try:
                        target = TargetInfo(
                            url=row.get(url_field, '').strip(),
                            username=row.get(username_field, '').strip(),
                            password=row.get(password_field, '').strip(),
                            index=len(targets),
                            source_file=Path(filepath).name
                        )
                        
                        if target.is_valid():
                            targets.append(target)
                        else:
                            warnings.append(f"第{row_num}行数据不完整")
                            
                    except Exception as e:
                        errors.append(f"第{row_num}行解析失败: {str(e)}")
            
            return ImportResult(
                success=len(targets) > 0,
                targets=targets,
                total_rows=total_rows,
                valid_count=len(targets),
                invalid_count=total_rows - len(targets),
                errors=errors,
                warnings=warnings,
                format=ImportFormat.TSV,
                encoding=encoding
            )
            
        except Exception as e:
            return ImportResult(
                success=False,
                targets=[],
                total_rows=0,
                valid_count=0,
                invalid_count=0,
                errors=[f"TSV导入失败: {str(e)}"],
                format=ImportFormat.TSV
            )
    
    def import_text(self, filepath: str) -> ImportResult:
        """导入文本文件"""
        targets: List[TargetInfo] = []
        errors: List[str] = []
        warnings: List[str] = []
        total_rows = 0
        
        try:
            encoding = self.detect_encoding(filepath)
            
            with open(filepath, 'r', encoding=encoding) as f:
                lines = f.readlines()
            
            for row_num, line in enumerate(lines, start=1):
                line = line.strip()
                if not line or line.startswith('#'):
                    continue
                
                total_rows += 1
                
                try:
                    parts = None
                    for sep in [',', '\t', '|', ' ']:
                        if sep in line:
                            parts = line.split(sep)
                            break
                    
                    if not parts or len(parts) < 3:
                        warnings.append(f"第{row_num}行格式错误，需要: URL,用户名,密码")
                        continue
                    
                    target = TargetInfo(
                        url=parts[0].strip(),
                        username=parts[1].strip(),
                        password=parts[2].strip(),
                        index=len(targets),
                        source_file=Path(filepath).name
                    )
                    
                    if target.is_valid():
                        targets.append(target)
                    else:
                        warnings.append(f"第{row_num}行数据不完整")
                        
                except Exception as e:
                    errors.append(f"第{row_num}行解析失败: {str(e)}")
            
            return ImportResult(
                success=len(targets) > 0,
                targets=targets,
                total_rows=total_rows,
                valid_count=len(targets),
                invalid_count=total_rows - len(targets),
                errors=errors,
                warnings=warnings,
                format=ImportFormat.TEXT,
                encoding=encoding
            )
            
        except Exception as e:
            return ImportResult(
                success=False,
                targets=[],
                total_rows=0,
                valid_count=0,
                invalid_count=0,
                errors=[f"文本文件导入失败: {str(e)}"],
                format=ImportFormat.TEXT
            )
    
    def import_json(self, filepath: str) -> ImportResult:
        """导入JSON文件"""
        targets: List[TargetInfo] = []
        errors: List[str] = []
        warnings: List[str] = []
        total_rows = 0
        
        try:
            encoding = self.detect_encoding(filepath)
            
            with open(filepath, 'r', encoding=encoding) as f:
                data = json.load(f)
            
            if isinstance(data, list):
                items = data
            elif isinstance(data, dict) and 'targets' in data:
                items = data['targets']
            else:
                return ImportResult(
                    success=False,
                    targets=[],
                    total_rows=0,
                    valid_count=0,
                    invalid_count=0,
                    errors=["JSON格式不支持，需要数组或包含targets字段的对象"]
                )
            
            for idx, item in enumerate(items):
                total_rows += 1
                
                try:
                    if not isinstance(item, dict):
                        warnings.append(f"第{idx}项不是对象，已跳过")
                        continue
                    
                    url = item.get('url', '').strip()
                    username = item.get('username', '').strip()
                    password = item.get('password', '').strip()
                    
                    extra: Dict[str, str] = {}
                    for key, value in item.items():
                        if key not in ['url', 'username', 'password'] and value:
                            extra[key] = str(value)
                    
                    target = TargetInfo(
                        url=url,
                        username=username,
                        password=password,
                        extra=extra if extra else None,
                        index=len(targets),
                        source_file=Path(filepath).name
                    )
                    
                    if target.is_valid():
                        targets.append(target)
                    else:
                        warnings.append(f"第{idx}项数据不完整，已跳过")
                        
                except Exception as e:
                    errors.append(f"第{idx}项解析失败: {str(e)}")
            
            return ImportResult(
                success=len(targets) > 0,
                targets=targets,
                total_rows=total_rows,
                valid_count=len(targets),
                invalid_count=total_rows - len(targets),
                errors=errors,
                warnings=warnings,
                format=ImportFormat.JSON,
                encoding=encoding
            )
            
        except Exception as e:
            return ImportResult(
                success=False,
                targets=[],
                total_rows=0,
                valid_count=0,
                invalid_count=0,
                errors=[f"JSON导入失败: {str(e)}"],
                format=ImportFormat.JSON
            )
    
    def _find_field(self, fields: List[str], candidates: Set[str]) -> Optional[str]:
        """查找匹配的字段名"""
        for candidate in candidates:
            for field in fields:
                if field.lower() == candidate.lower():
                    return field
        return None
    
    def _find_field_index(self, fields: List[str], candidates: Set[str]) -> Optional[int]:
        """查找匹配的字段索引"""
        for candidate in candidates:
            for i, field in enumerate(fields):
                if field and field.lower() == candidate.lower():
                    return i
        return None
    
    def export_template(
        self,
        filepath: str,
        format: ImportFormat = ImportFormat.CSV,
        sample_count: int = 3
    ) -> bool:
        """
        导出模板文件
        
        Args:
            filepath: 文件路径
            format: 导出格式
            sample_count: 示例数量
            
        Returns:
            是否成功
        """
        try:
            if format == ImportFormat.CSV:
                with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f)
                    writer.writerow(['URL', '用户名', '密码'])
                    for i in range(sample_count):
                        writer.writerow([
                            f'https://example{i+1}.com/login',
                            f'user{i+1}',
                            f'password{i+1}'
                        ])
                return True
                
            elif format == ImportFormat.EXCEL and self.openpyxl_available:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "目标列表"
                ws.append(['URL', '用户名', '密码'])
                for i in range(sample_count):
                    ws.append([
                        f'https://example{i+1}.com/login',
                        f'user{i+1}',
                        f'password{i+1}'
                    ])
                wb.save(filepath)
                wb.close()
                return True
                
            elif format == ImportFormat.JSON:
                data = {
                    'targets': [
                        {
                            'url': f'https://example{i+1}.com/login',
                            'username': f'user{i+1}',
                            'password': f'password{i+1}'
                        }
                        for i in range(sample_count)
                    ]
                }
                with open(filepath, 'w', encoding='utf-8') as f:
                    json.dump(data, f, indent=2, ensure_ascii=False)
                return True
                
            else:
                return False
                
        except Exception as e:
            print(f"导出模板失败: {e}")
            return False


def get_supported_formats() -> List[str]:
    """获取支持的导入格式"""
    formats = ['.csv', '.tsv', '.txt', '.json']
    
    if OPENPYXL_AVAILABLE or PANDAS_AVAILABLE:
        formats.extend(['.xlsx', '.xls'])
    
    return formats


if __name__ == "__main__":
    print("支持的导入格式:", get_supported_formats())
    
    importer = EnhancedBatchImporter()
    
    test_csv = "test_import.csv"
    with open(test_csv, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(['URL', '用户名', '密码'])
        writer.writerow(['https://httpbin.org/post', 'testuser1', 'testpass1'])
        writer.writerow(['https://httpbin.org/post', 'testuser2', 'testpass2'])
    
    result = importer.import_csv(test_csv)
    print(f"\n导入结果:")
    print(f"  成功: {result.success}")
    print(f"  总行数: {result.total_rows}")
    print(f"  有效数: {result.valid_count}")
    print(f"  无效数: {result.invalid_count}")
    print(f"  编码: {result.encoding}")
    print(f"  目标列表:")
    for target in result.targets:
        print(f"    - {target.url} | {target.username}")
    
    os.remove(test_csv)
